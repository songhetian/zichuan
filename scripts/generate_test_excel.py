#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成多行测试Excel文件
模拟多个员工的设备信息
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 模拟的硬件信息（使用田鹤松的真实硬件）
tian_hesong_components = {
    "CPU": {"name": "12th Gen Intel Core i5-12400", "brand": "GenuineIntel"},
    "内存": {"name": "16GB DDR3200MHz", "brand": "Shenzhen Zhongshi Technology Co Ltd"},
    "硬盘": {"name": "953GB HDD (Dahua)", "brand": "未知"},
    "主板": {"name": "H610M-D", "brand": "Colorful Technology And Development Co.,LTD"},
    "显卡": {"name": "GameViewer Virtual Display Adapter (1GB)", "brand": "未知"},
}

# 测试数据 - 多行
test_data = [
    {
        "使用人": "田鹤松",
        "部门": "后勤部",
        "设备名称": "田鹤松的电脑主机",
        "设备分类": "电脑主机",
        "设备分类编号": "PC",
        **tian_hesong_components,
    },
    {
        "使用人": "张三",
        "部门": "技术部",
        "设备名称": "张三的电脑主机",
        "设备分类": "电脑主机",
        "设备分类编号": "PC",
        "CPU": {"name": "Intel Core i7-12700", "brand": "Intel"},
        "内存": {"name": "32GB DDR4 3200MHz", "brand": "Kingston"},
        "硬盘": {"name": "1TB NVMe SSD", "brand": "Samsung"},
        "主板": {"name": "B660M", "brand": "ASUS"},
        "显卡": {"name": "NVIDIA RTX 3060", "brand": "NVIDIA"},
    },
    {
        "使用人": "李四",
        "部门": "技术部",
        "设备名称": "李四的电脑主机",
        "设备分类": "电脑主机",
        "设备分类编号": "PC",
        "CPU": {"name": "Intel Core i5-12400", "brand": "Intel"},
        "内存": {"name": "16GB DDR4 3200MHz", "brand": "Kingston"},
        "硬盘": {"name": "512GB NVMe SSD", "brand": "Samsung"},
        "主板": {"name": "B660M", "brand": "ASUS"},
        "显卡": {"name": "NVIDIA GTX 1650", "brand": "NVIDIA"},
    },
    {
        "使用人": "王五",
        "部门": "财务部",
        "设备名称": "王五的电脑主机",
        "设备分类": "电脑主机",
        "设备分类编号": "PC",
        "CPU": {"name": "Intel Core i7-12700", "brand": "Intel"},
        "内存": {"name": "32GB DDR4 3200MHz", "brand": "Kingston"},
        "硬盘": {"name": "1TB NVMe SSD", "brand": "Samsung"},
        "主板": {"name": "B660M", "brand": "ASUS"},
        "显卡": {"name": "NVIDIA RTX 3060", "brand": "NVIDIA"},
    },
]

headers = [
    "使用人",
    "部门",
    "设备名称",
    "设备分类",
    "设备分类编号",
    "CPU型号",
    "CPU品牌",
    "内存型号",
    "内存品牌",
    "硬盘型号",
    "硬盘品牌",
    "主板型号",
    "主板品牌",
    "显卡型号",
    "显卡品牌",
]

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_alignment = Alignment(vertical="center", wrap_text=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "设备导入"

# 写表头
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 写数据行
for row_idx, row_data in enumerate(test_data, 2):
    values = [
        row_data["使用人"],
        row_data["部门"],
        row_data["设备名称"],
        row_data["设备分类"],
        row_data["设备分类编号"],
        row_data["CPU"]["name"],
        row_data["CPU"]["brand"],
        row_data["内存"]["name"],
        row_data["内存"]["brand"],
        row_data["硬盘"]["name"],
        row_data["硬盘"]["brand"],
        row_data["主板"]["name"],
        row_data["主板"]["brand"],
        row_data["显卡"]["name"],
        row_data["显卡"]["brand"],
    ]
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = data_alignment
        cell.border = thin_border

# 设置列宽
col_widths = [15, 15, 25, 10, 12, 35, 15, 25, 15, 25, 15, 20, 15, 30, 15]
for col, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

# 设置行高
ws.row_dimensions[1].height = 25
for row_idx in range(2, len(test_data) + 2):
    ws.row_dimensions[row_idx].height = 25

output_path = "测试导入_多行数据.xlsx"
wb.save(output_path)
print(f"测试文件已生成: {output_path}")
print(f"共 {len(test_data)} 行数据")
