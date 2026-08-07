#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
硬件信息扫描脚本 - 资产管理系统配套工具
扫描Windows电脑硬件信息，生成Excel导入文件

使用方法:
    1. 运行脚本: python hardware_scanner.py
    2. 输入使用人姓名和部门
    3. 脚本生成 asset_import.xlsx 文件

打包成exe:
    pyinstaller --onefile --windowed --name 硬件扫描工具 hardware_scanner.py
"""

import json
import argparse
import sys
import subprocess
import re
from datetime import datetime
from pathlib import Path

from hardware_utils import (
    parse_memory_output,
    parse_memory_ps_output,
    pick_memory_result,
    guess_disk_brand,
    map_monitor_brand,
    parse_cpu_ps_output,
    parse_motherboard_ps_output,
    parse_gpu_ps_output,
)

# 检测是否在无控制台窗口模式下运行（打包成 .exe 后 console=False）
def has_console():
    try:
        sys.stdin.fileno()
        return True
    except (OSError, AttributeError, RuntimeError):
        return False

# 无控制台时使用 tkinter 弹窗输入
def gui_input(title, prompt, required=True):
    import tkinter as tk
    from tkinter import simpledialog, messagebox

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    while True:
        value = simpledialog.askstring(title, prompt, parent=root)
        if value is None:
            root.destroy()
            sys.exit(0)
        value = value.strip()
        if not required or value:
            root.destroy()
            return value
        messagebox.showwarning("输入错误", "此项为必填项，请重新输入。", parent=root)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("正在安装依赖...")
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter


def run_wmic(query, timeout=8):
    """执行 wmic 命令。带超时防止 wmic 偶发卡死导致整机扫描挂起。"""
    try:
        result = subprocess.run(
            ["wmic"] + query.split(),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="ignore",
            timeout=timeout
        )
        return result.stdout.strip()
    except subprocess.TimeoutExpired:
        print(f"WMIC命令超时(>{timeout}s): {query}")
        return ""
    except Exception as e:
        print(f"WMIC命令失败: {e}")
        return ""


def run_ps(script, timeout=10):
    """执行 PowerShell 命令，返回 stdout（wmic fallback 用）。"""
    try:
        result = subprocess.run(
            ["powershell", "-Command", script],
            capture_output=True, text=True, encoding="utf-8", errors="ignore", timeout=timeout
        )
        return result.stdout.strip()
    except Exception:
        return ""


def get_cpu_info():
    # 方法1：wmic
    output = run_wmic("cpu get Name,Manufacturer /value")
    name = ""
    brand = ""
    for line in output.split("\n"):
        line = line.strip()
        if line.startswith("Name="):
            name = line.replace("Name=", "").strip()
        elif line.startswith("Manufacturer="):
            brand = line.replace("Manufacturer=", "").strip()

    # 方法2：PowerShell fallback（wmic 失败/卡死时）
    if not name:
        ps = run_ps(
            "Get-CimInstance -ClassName Win32_Processor -ErrorAction SilentlyContinue | "
            "ForEach-Object { Write-Output \"$($_.Name)|$($_.Manufacturer)\" }"
        )
        parsed = parse_cpu_ps_output(ps)
        if parsed:
            return parsed

    name = re.sub(r"\s+", " ", name)
    name = name.replace("(R)", "").replace("(TM)", "").replace("  ", " ")

    if not name:
        name = "未知CPU"
    if not brand:
        brand = "未知"

    return {"category": "CPU", "name": name, "brand": brand}


def get_memory_info_list():
    """扫描所有内存条，返回列表（每根内存条单独一条记录）

    优先 wmic memorychip，失败/空结果时回退到 PowerShell Get-CimInstance
    （wmic 在 Windows 11 24H2+ 已被移除）。
    去重：按插槽(DeviceLocator)去重，修复部分主板 wmic 重复报告导致"2条变3条"；
    无插槽信息时保守保留全部（避免"2条变1条"漏报）。
    """
    modules = []

    # 方法1：wmic
    output = run_wmic("memorychip get DeviceLocator,Capacity,Speed,Manufacturer /value")
    wmic_modules = parse_memory_output(output) if output else []

    # 方法2：PowerShell Get-CimInstance（wmic 缺失或结果无插槽信息时用于交叉验证）
    ps_modules = []
    try:
        ps_script = (
            "Get-CimInstance -ClassName Win32_PhysicalMemory -ErrorAction SilentlyContinue | "
            "ForEach-Object { Write-Output \"$($_.DeviceLocator)|$($_.Capacity)|$($_.Speed)|$($_.Manufacturer)\" }"
        )
        result = subprocess.run(
            ["powershell", "-Command", ps_script],
            capture_output=True, text=True, encoding="utf-8", errors="ignore", timeout=15
        )
        if result.stdout.strip():
            ps_modules = parse_memory_ps_output(result.stdout)
    except Exception:
        pass

    # 选择更可靠的结果：wmic 有插槽用 wmic；否则用 PS（PS 的 DeviceLocator 通常可靠）
    modules = pick_memory_result(wmic_modules, ps_modules)

    if not modules:
        return [{"category": "内存", "name": "未知内存", "brand": "未知"}]

    return modules


def _guess_disk_brand(model):
    """从硬盘型号推断品牌（使用 hardware_utils 的扩充关键词表）"""
    return guess_disk_brand(model)


def get_disk_info_list():
    """扫描所有物理硬盘，返回列表（每块硬盘单独一条记录）

    使用 PowerShell Get-CimInstance（最可靠），fallback 到 wmic
    """
    disks = []

    # 方法1：PowerShell + Get-CimInstance（推荐，格式可靠）
    try:
        ps_script = (
            "Get-CimInstance -ClassName Win32_DiskDrive -ErrorAction SilentlyContinue | "
            "Where-Object { $_.MediaType -notlike '*USB*' -and $_.Model -notlike '*Virtual*' -and $_.Size -gt 0 } | "
            "ForEach-Object { "
            "Write-Output \"$($_.Model)|$($_.Size)|$($_.MediaType)\" "
            "}"
        )
        result = subprocess.run(
            ["powershell", "-Command", ps_script],
            capture_output=True, text=True, encoding="utf-8", errors="ignore", timeout=15
        )
        for line in result.stdout.strip().split("\n"):
            line = line.strip()
            if not line:
                continue
            parts = line.split("|")
            if len(parts) >= 2:
                model = parts[0].strip()
                size_str = parts[1].strip()
                media_type = parts[2].strip() if len(parts) > 2 else ""
                if model and size_str and size_str.isdigit():
                    size_gb = int(size_str) // (1024 ** 3)
                    if size_gb > 0:
                        disks.append({"model": model, "size": size_gb, "media_type": media_type})
    except Exception:
        pass

    # 方法2：wmic fallback（旧系统兼容）
    if not disks:
        try:
            output = run_wmic("diskdrive get Model,Size,MediaType /value")
            # wmic /value 格式不稳定，改用 list brief 格式
            output_brief = run_wmic("diskdrive get Model,Size,MediaType /format:list")
            # 尝试解析 list 格式
            current = {}
            for line in output_brief.split("\n"):
                line = line.strip()
                if not line:
                    if current and current.get("model") and current.get("size", 0) > 0:
                        disks.append(current)
                    current = {}
                    continue
                if "=" in line:
                    key, val = line.split("=", 1)
                    key = key.strip()
                    val = val.strip()
                    if key == "Model":
                        current["model"] = val
                    elif key == "Size" and val.isdigit():
                        current["size"] = int(val) // (1024 ** 3)
                    elif key == "MediaType":
                        current["media_type"] = val
            if current and current.get("model") and current.get("size", 0) > 0:
                disks.append(current)
        except Exception:
            pass

    # 过滤 USB 和虚拟设备
    physical_disks = [
        d for d in disks
        if "USB" not in d.get("model", "") and "Virtual" not in d.get("model", "")
    ]

    if not physical_disks:
        return [{"category": "硬盘", "name": "未知硬盘", "brand": "未知"}]

    result = []
    for disk in physical_disks:
        model = disk.get("model", "")
        size_gb = disk.get("size", 0)
        brand = _guess_disk_brand(model)

        media_type = disk.get("media_type", "")
        disk_type = "SSD" if ("SSD" in media_type or "Solid" in media_type or "SSD" in model) else "HDD"

        if size_gb >= 1024:
            size_str = f"{size_gb / 1024:.0f}TB".replace(".0", "")
        else:
            size_str = f"{size_gb}GB"

        name = f"{size_str} {disk_type}"
        if model:
            short_model = model.split()[0] if len(model.split()) > 1 else model
            name = f"{size_str} {disk_type} ({short_model})"

        result.append({"category": "硬盘", "name": name, "brand": brand})

    return result


def get_motherboard_info():
    output = run_wmic("baseboard get Product,Manufacturer /value")
    name = ""
    brand = ""

    for line in output.split("\n"):
        line = line.strip()
        if line.startswith("Product="):
            name = line.replace("Product=", "").strip()
        elif line.startswith("Manufacturer="):
            brand = line.replace("Manufacturer=", "").strip()

    # 方法2：PowerShell fallback（wmic 失败/卡死时）
    if not name or name in ["Base Board", "Not Available"]:
        ps = run_ps(
            "Get-CimInstance -ClassName Win32_BaseBoard -ErrorAction SilentlyContinue | "
            "ForEach-Object { Write-Output \"$($_.Product)|$($_.Manufacturer)\" }"
        )
        parsed = parse_motherboard_ps_output(ps)
        if parsed:
            return parsed

    if not name or name in ["Base Board", "Not Available"]:
        name = "未知主板"
    if not brand:
        brand = "未知"

    return {"category": "主板", "name": name, "brand": brand}


def get_gpu_info():
    output = run_wmic("path win32_VideoController get Name,AdapterRAM /value")
    name = ""
    vram = ""

    current = {}
    for line in output.split("\n"):
        line = line.strip()
        if line.startswith("Name="):
            if current and current.get("name"):
                break
            current["name"] = line.replace("Name=", "").strip()
        elif line.startswith("AdapterRAM=") and current:
            val = line.replace("AdapterRAM=", "").strip()
            if val and val.isdigit():
                vram_gb = int(val) // (1024 ** 3)
                if vram_gb > 0:
                    vram = f"{vram_gb}GB"

    name = current.get("name", "")

    # 方法2：PowerShell fallback（wmic 失败/卡死时）
    if not name:
        ps = run_ps(
            "Get-CimInstance -ClassName Win32_VideoController -ErrorAction SilentlyContinue | "
            "ForEach-Object { Write-Output \"$($_.Name)|$($_.AdapterRAM)\" }"
        )
        parsed = parse_gpu_ps_output(ps)
        if parsed:
            return parsed

    if not name:
        return None

    brand = "未知"
    if "NVIDIA" in name or "GeForce" in name or "RTX" in name or "GTX" in name:
        brand = "NVIDIA"
    elif "AMD" in name or "Radeon" in name:
        brand = "AMD"
    elif "Intel" in name:
        brand = "Intel"

    display_name = name
    if vram:
        display_name = f"{name} ({vram})"

    return {"category": "显卡", "name": display_name, "brand": brand}


def is_laptop():
    """检测当前设备是否是笔记本电脑

    判断依据（满足任一即认为是笔记本）：
    1. 存在电池（Win32_Battery）
    2. 机箱类型为便携式（ChassisType: 8=Portable, 9=Laptop, 10=Notebook, 14=SubNotebook）
    """
    # 方法1：检查电池（最可靠）
    try:
        result = subprocess.run(
            ["powershell", "-Command",
             "Get-CimInstance -ClassName Win32_Battery -ErrorAction SilentlyContinue | Select-Object -First 1 | Measure-Object | Select-Object -ExpandProperty Count"],
            capture_output=True, text=True, encoding="utf-8", errors="ignore", timeout=10
        )
        if result.stdout.strip() and int(result.stdout.strip()) > 0:
            return True
    except Exception:
        pass

    # 方法2：检查机箱类型
    try:
        output = run_wmic("systemenclosure get ChassisType /value")
        for line in output.split("\n"):
            line = line.strip()
            if line.startswith("ChassisType="):
                # ChassisType 可能是 "{8}" 或 "8" 格式
                types = re.findall(r'\d+', line.replace("ChassisType=", ""))
                # 8=Portable, 9=Laptop, 10=Notebook, 14=SubNotebook
                for t in types:
                    if t in ["8", "9", "10", "14"]:
                        return True
    except Exception:
        pass

    return False


def get_monitor_info_list():
    """扫描所有显示器，返回列表（每台显示器单独一条记录）

    使用 PowerShell + WmiMonitorID（最可靠），fallback 到 wmic desktopmonitor
    """
    monitors = []

    # 方法1：PowerShell + WmiMonitorID（推荐，支持 Windows 10/11）
    try:
        ps_script = (
            "Get-CimInstance -Namespace root\\wmi -ClassName WmiMonitorID -ErrorAction SilentlyContinue | "
            "ForEach-Object { "
            "$name = ($_.UserFriendlyName | Where-Object {$_ -ne 0} | ForEach-Object {[char]$_}) -join ''; "
            "$mfr = ($_.ManufacturerName | Where-Object {$_ -ne 0} | ForEach-Object {[char]$_}) -join ''; "
            "Write-Output \"$name|$mfr\" "
            "}"
        )
        result = subprocess.run(
            ["powershell", "-Command", ps_script],
            capture_output=True, text=True, encoding="utf-8", errors="ignore", timeout=15
        )
        for line in result.stdout.strip().split("\n"):
            line = line.strip()
            if not line:
                continue
            parts = line.split("|", 1)
            name = parts[0].strip() if parts else ""
            brand = parts[1].strip() if len(parts) > 1 else ""

            if not name:
                name = "未知显示器"
            if not brand:
                brand = "未知"
            else:
                # EDID 三字母厂商码 -> 品牌全名（如 DEL -> Dell）
                mapped = map_monitor_brand(brand)
                if mapped:
                    brand = mapped

            monitors.append({"category": "显示器", "name": name, "brand": brand})
    except Exception:
        pass

    # 方法2：fallback 到 wmic desktopmonitor（旧系统兼容）
    if not monitors:
        try:
            output = run_wmic("desktopmonitor get Name,MonitorType /value")
            current = {}
            for line in output.split("\n"):
                line = line.strip()
                if not line:
                    if current:
                        name = current.get("name", "")
                        if name and name not in ["Default Monitor", "Generic PnP Monitor"]:
                            monitors.append({"category": "显示器", "name": name, "brand": "未知"})
                        current = {}
                    continue
                if line.startswith("Name="):
                    val = line.replace("Name=", "").strip()
                    if val:
                        current["name"] = val
                elif line.startswith("MonitorType="):
                    current["type"] = line.replace("MonitorType=", "").strip()
            if current and current.get("name"):
                monitors.append({"category": "显示器", "name": current["name"], "brand": "未知"})
        except Exception:
            pass

    return monitors


def scan_hardware():
    print("正在扫描硬件信息...")

    # 先检测设备类型（笔记本/台式机）
    laptop = is_laptop()
    if laptop:
        print("  设备类型: 笔记本电脑")
    else:
        print("  设备类型: 台式机")

    components = []

    cpu = get_cpu_info()
    if cpu:
        components.append(cpu)
        print(f"  CPU: {cpu['name']} ({cpu['brand']})")

    memory_list = get_memory_info_list()
    for i, mem in enumerate(memory_list, 1):
        components.append(mem)
        print(f"  内存{i}: {mem['name']} ({mem['brand']})")

    disk_list = get_disk_info_list()
    for i, disk in enumerate(disk_list, 1):
        components.append(disk)
        print(f"  硬盘{i}: {disk['name']} ({disk['brand']})")

    motherboard = get_motherboard_info()
    if motherboard:
        components.append(motherboard)
        print(f"  主板: {motherboard['name']} ({motherboard['brand']})")

    gpu = get_gpu_info()
    if gpu:
        components.append(gpu)
        print(f"  显卡: {gpu['name']} ({gpu['brand']})")

    # 显示器扫描
    monitor_list = get_monitor_info_list()
    for i, mon in enumerate(monitor_list, 1):
        components.append(mon)
        print(f"  显示器{i}: {mon['name']} ({mon['brand']})")

    return components, laptop


def generate_excel(employee_name, department_name, device_name=None, components=None, output_path=None, is_laptop_device=False):
    if components is None:
        components, is_laptop_device = scan_hardware()

    # 根据设备类型决定分类和编号
    if is_laptop_device:
        category_name = "笔记本电脑"
        category_code = "NB"
        default_device_name = f"{employee_name}的笔记本电脑"
    else:
        category_name = "电脑主机"
        category_code = "PC"
        default_device_name = f"{employee_name}的电脑主机"

    if not device_name:
        device_name = default_device_name

    # 按类别分组配件（内存/硬盘/显示器可能有多个）
    cpu_list = [c for c in components if c["category"] == "CPU"]
    memory_list = [c for c in components if c["category"] == "内存"]
    disk_list = [c for c in components if c["category"] == "硬盘"]
    motherboard_list = [c for c in components if c["category"] == "主板"]
    gpu_list = [c for c in components if c["category"] == "显卡"]
    monitor_list = [c for c in components if c["category"] == "显示器"]

    # 动态构建表头和数据行
    headers = [
        "使用人",
        "部门",
        "设备名称",
        "设备分类",
        "设备分类编号",
        "CPU型号",
        "CPU品牌",
    ]

    row_data = [
        employee_name,
        department_name,
        device_name,
        category_name,
        category_code,
        cpu_list[0]["name"] if cpu_list else "",
        cpu_list[0]["brand"] if cpu_list else "",
    ]

    # 内存：单列格式 + 多列格式（只有当有多个内存时才生成多列）
    headers.append("内存型号")
    headers.append("内存品牌")
    row_data.append(memory_list[0]["name"] if memory_list else "")
    row_data.append(memory_list[0]["brand"] if memory_list else "")

    # 只有当内存数量 > 1 时，才生成多列格式
    if len(memory_list) > 1:
        for i in range(len(memory_list)):
            idx = i + 1
            headers.append(f"内存{idx}型号")
            headers.append(f"内存{idx}品牌")
            row_data.append(memory_list[i]["name"])
            row_data.append(memory_list[i]["brand"])

    # 硬盘：单列格式 + 多列格式（只有当有多个硬盘时才生成多列）
    headers.append("硬盘型号")
    headers.append("硬盘品牌")
    row_data.append(disk_list[0]["name"] if disk_list else "")
    row_data.append(disk_list[0]["brand"] if disk_list else "")

    # 只有当硬盘数量 > 1 时，才生成多列格式
    if len(disk_list) > 1:
        for i in range(len(disk_list)):
            idx = i + 1
            headers.append(f"硬盘{idx}型号")
            headers.append(f"硬盘{idx}品牌")
            row_data.append(disk_list[i]["name"])
            row_data.append(disk_list[i]["brand"])

    # 主板
    headers.append("主板型号")
    headers.append("主板品牌")
    row_data.append(motherboard_list[0]["name"] if motherboard_list else "")
    row_data.append(motherboard_list[0]["brand"] if motherboard_list else "")

    # 显卡
    headers.append("显卡型号")
    headers.append("显卡品牌")
    row_data.append(gpu_list[0]["name"] if gpu_list else "")
    row_data.append(gpu_list[0]["brand"] if gpu_list else "")

    # 显示器：单列格式 + 多列格式（只有当有多个显示器时才生成多列）
    headers.append("显示器型号")
    headers.append("显示器品牌")
    row_data.append(monitor_list[0]["name"] if monitor_list else "")
    row_data.append(monitor_list[0]["brand"] if monitor_list else "")

    # 只有当显示器数量 > 1 时，才生成多列格式
    if len(monitor_list) > 1:
        for i in range(len(monitor_list)):
            idx = i + 1
            headers.append(f"显示器{idx}型号")
            headers.append(f"显示器{idx}品牌")
            row_data.append(monitor_list[i]["name"])
            row_data.append(monitor_list[i]["brand"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "设备导入"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    data_alignment = Alignment(vertical="center", wrap_text=True)
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.alignment = data_alignment
        cell.border = thin_border

    # 设置列宽
    for col in range(1, len(headers) + 1):
        header_name = headers[col - 1]
        if "型号" in header_name:
            width = 30 if "CPU" in header_name or "显卡" in header_name else 25
        elif "品牌" in header_name:
            width = 15
        elif header_name in ("使用人", "部门"):
            width = 15
        elif header_name == "设备名称":
            width = 25
        elif header_name in ("设备分类",):
            width = 10
        elif header_name == "设备分类编号":
            width = 12
        else:
            width = 18
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25

    final_path = Path(output_path) if output_path else Path("asset_import.xlsx")

    # 安全保存：如果文件被占用（如 Excel 正在打开），自动使用带时间戳的新文件名
    def safe_save_workbook(wb, path):
        try:
            wb.save(path)
            return path, None
        except PermissionError:
            # 文件被占用，生成带时间戳的新文件名
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_path = path.with_name(f"{path.stem}_{ts}{path.suffix}")
            try:
                wb.save(new_path)
                return new_path, path.name
            except Exception as e:
                raise e

    saved_path, original_name = safe_save_workbook(wb, final_path)
    print(f"\n已保存到: {saved_path.absolute()}")
    if original_name:
        print(f"提示: {original_name} 正在被占用（可能 Excel 正在打开），已自动保存为 {saved_path.name}")
    return saved_path


def main():
    parser = argparse.ArgumentParser(description="扫描电脑硬件信息并生成资产导入Excel")
    parser.add_argument("--employee", "-e", help="使用人姓名")
    parser.add_argument("--department", "-d", help="所属部门")
    parser.add_argument("--name", "-n", help="设备名称（默认: 姓名+的电脑主机/笔记本电脑）")
    parser.add_argument("--output", "-o", default="asset_import.xlsx", help="输出Excel文件路径")
    parser.add_argument("--scan-only", action="store_true", help="仅扫描硬件，不指定使用人")

    args = parser.parse_args()

    if not args.scan_only and not (args.employee and args.department):
        if has_console():
            print("=" * 50)
            print("          资产管理系统 - 硬件扫描工具")
            print("=" * 50)
            print()
            args.employee = input("请输入使用人姓名: ").strip()
            while not args.employee:
                args.employee = input("姓名不能为空，请重新输入: ").strip()

            args.department = input("请输入所属部门: ").strip()
            while not args.department:
                args.department = input("部门不能为空，请重新输入: ").strip()

            print()
        else:
            args.employee = gui_input("硬件扫描工具", "请输入使用人姓名：")
            args.department = gui_input("硬件扫描工具", "请输入所属部门：")

    components, laptop = scan_hardware()

    if not components:
        print("错误: 未能扫描到任何硬件信息")
        sys.exit(1)

    if args.scan_only:
        print("\n硬件信息扫描完成，未生成导入文件。")
        sys.exit(0)

    saved_path = generate_excel(
        employee_name=args.employee,
        department_name=args.department,
        device_name=args.name,
        components=components,
        output_path=args.output,
        is_laptop_device=laptop
    )

    file_name = saved_path.name

    if has_console():
        device_type = "笔记本电脑" if laptop else "电脑主机"
        print("\n" + "=" * 50)
        print("          扫描完成！")
        print("=" * 50)
        print(f" 使用人: {args.employee}")
        print(f" 部门: {args.department}")
        print(f" 设备类型: {device_type}")
        print(f" 文件: {file_name}")
        print("\n请将此文件发送给管理员进行导入。")
        print("=" * 50)

        input("\n按回车键退出...")
    else:
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        device_type = "笔记本电脑" if laptop else "电脑主机"
        messagebox.showinfo(
            "扫描完成",
            f"使用人: {args.employee}\n"
            f"部门: {args.department}\n"
            f"设备类型: {device_type}\n"
            f"文件: {file_name}\n\n"
            "请将此文件发送给管理员进行导入。",
            parent=root
        )
        root.destroy()


if __name__ == "__main__":
    main()
