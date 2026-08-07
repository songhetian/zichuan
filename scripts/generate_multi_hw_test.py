#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成多内存、多硬盘的测试Excel文件
用于验证硬件扫描导入功能
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "设备导入"

headers = [
    "使用人",
    "部门",
    "设备名称",
    "设备分类",
    "设备分类编号",
    "CPU型号",
    "CPU品牌",
    # 单列兼容格式
    "内存型号",
    "内存品牌",
    # 多列格式
    "内存1型号",
    "内存1品牌",
    "内存2型号",
    "内存2品牌",
    # 单列兼容格式
    "硬盘型号",
    "硬盘品牌",
    # 多列格式
    "硬盘1型号",
    "硬盘1品牌",
    "硬盘2型号",
    "硬盘2品牌",
    "主板型号",
    "主板品牌",
    "显卡型号",
    "显卡品牌",
]

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

test_data = [
    # 测试用例1：两条不同品牌的8GB内存 + SSD + HDD 双硬盘
    {
        "使用人": "张三",
        "部门": "技术部",
        "设备名称": "张三的高配电脑",
        "设备分类": "电脑主机",
        "设备分类编号": "PC",
        "CPU型号": "Intel Core i7-12700",
        "CPU品牌": "Intel",
        "内存型号": "8GB DDR4 3200MHz",
        "内存品牌": "Kingston",
        "内存1型号": "8GB DDR4 3200MHz",
        "内存1品牌": "Kingston",
        "内存2型号": "8GB DDR4 3200MHz",
        "内存2品牌": "ADATA",
        "硬盘型号": "512GB SSD (SAMSUNG)",
        "硬盘品牌": "Samsung",
        "硬盘1型号": "512GB SSD (SAMSUNG)",
        "硬盘1品牌": "Samsung",
        "硬盘2型号": "2TB HDD (WD)",
        "硬盘2品牌": "Western Digital",
        "主板型号": "B660M",
        "主板品牌": "ASUS",
        "显卡型号": "NVIDIA RTX 3060 (12GB)",
        "显卡品牌": "NVIDIA",
    },
    # 测试用例2：两条相同品牌的16GB内存 + 单块1TB SSD
    {
        "使用人": "李四",
        "部门": "设计部",
        "设备名称": "李四的设计工作站",
        "设备分类": "电脑主机",
        "设备分类编号": "PC",
        "CPU型号": "Intel Core i9-13900",
        "CPU品牌": "Intel",
        "内存型号": "16GB DDR5 4800MHz",
        "内存品牌": "Kingston",
        "内存1型号": "16GB DDR5 4800MHz",
        "内存1品牌": "Kingston",
        "内存2型号": "16GB DDR5 4800MHz",
        "内存2品牌": "Kingston",
        "硬盘型号": "1TB SSD (Samsung)",
        "硬盘品牌": "Samsung",
        "硬盘1型号": "1TB SSD (Samsung)",
        "硬盘1品牌": "Samsung",
        "硬盘2型号": "",
        "硬盘2品牌": "",
        "主板型号": "Z790",
        "主板品牌": "MSI",
        "显卡型号": "NVIDIA RTX 4070 (12GB)",
        "显卡品牌": "NVIDIA",
    },
    # 测试用例3：单根16GB内存 + 三块硬盘（SSD + HDD + HDD）
    {
        "使用人": "王五",
        "部门": "财务部",
        "设备名称": "王五的办公电脑",
        "设备分类": "电脑主机",
        "设备分类编号": "PC",
        "CPU型号": "Intel Core i5-12400",
        "CPU品牌": "Intel",
        "内存型号": "16GB DDR4 3200MHz",
        "内存品牌": "Kingston",
        "内存1型号": "16GB DDR4 3200MHz",
        "内存1品牌": "Kingston",
        "内存2型号": "",
        "内存2品牌": "",
        "硬盘型号": "256GB SSD (Kingston)",
        "硬盘品牌": "Kingston",
        "硬盘1型号": "256GB SSD (Kingston)",
        "硬盘1品牌": "Kingston",
        "硬盘2型号": "1TB HDD (Seagate)",
        "硬盘2品牌": "Seagate",
        "主板型号": "H610",
        "主板品牌": "ASUS",
        "显卡型号": "",
        "显卡品牌": "",
    },
]

data_alignment = Alignment(vertical="center", wrap_text=True)
for row_idx, data in enumerate(test_data, 2):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=data.get(header, ""))
        cell.alignment = data_alignment
        cell.border = thin_border

col_widths = []
for header in headers:
    if "型号" in header:
        if "CPU" in header or "显卡" in header:
            col_widths.append(35)
        else:
            col_widths.append(28)
    elif "品牌" in header:
        col_widths.append(18)
    elif header in ("使用人", "部门"):
        col_widths.append(12)
    elif header == "设备名称":
        col_widths.append(22)
    elif header == "设备分类":
        col_widths.append(10)
    elif header == "设备分类编号":
        col_widths.append(12)
    else:
        col_widths.append(15)

for col, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

ws.row_dimensions[1].height = 28

output_path = r"e:\System\zichuan\scripts\测试导入_多内存多硬盘.xlsx"
wb.save(output_path)
print(f"测试文件已生成: {output_path}")
print(f"共 {len(test_data)} 条测试数据:")
for i, data in enumerate(test_data, 1):
    print(f"  {i}. {data['使用人']} - {data['设备名称']}")
    print(f"     内存: {data.get('内存1型号', '')}({data.get('内存1品牌', '')}) + {data.get('内存2型号', '')}({data.get('内存2品牌', '')})")
    print(f"     硬盘: {data.get('硬盘1型号', '')} + {data.get('硬盘2型号', '')}")
