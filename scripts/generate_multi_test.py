#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成多人测试Excel文件
模拟不同部门、不同配置的电脑主机
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 模拟不同配置的电脑
test_data = [
    {
        "使用人": "张三",
        "部门": "技术部",
        "设备名称": "张三的电脑主机",
        "设备分类": "电脑主机",
        "设备分类编号": "PC",
        "CPU型号": "Intel Core i7-13700K",
        "CPU品牌": "Intel",
        "内存型号": "32GB DDR4 3600MHz",
        "内存品牌": "Corsair",
        "硬盘型号": "1TB NVMe SSD",
        "硬盘品牌": "Samsung",
        "主板型号": "Z790-A",
        "主板品牌": "ASUS",
        "显卡型号": "NVIDIA RTX 4070 Ti",
        "显卡品牌": "NVIDIA",
    },
    {
        "使用人": "李四",
        "部门": "技术部",
        "设备名称": "李四的电脑主机",
        "设备分类": "电脑主机",
        "设备分类编号": "PC",
        "CPU型号": "Intel Core i5-12400F",
        "CPU品牌": "Intel",
        "内存型号": "16GB DDR4 3200MHz",
        "内存品牌": "Kingston",
        "硬盘型号": "512GB NVMe SSD",
        "硬盘品牌": "Western Digital",
        "主板型号": "B660M",
        "主板品牌": "MSI",
        "显卡型号": "NVIDIA GTX 1660 Super",
        "显卡品牌": "NVIDIA",
    },
    {
        "使用人": "王五",
        "部门": "财务部",
        "设备名称": "王五的电脑主机",
        "设备分类": "电脑主机",
        "设备分类编号": "PC",
        "CPU型号": "Intel Core i5-10400",
        "CPU品牌": "Intel",
        "内存型号": "16GB DDR4 2666MHz",
        "内存品牌": "Kingston",
        "硬盘型号": "256GB SSD",
        "硬盘品牌": "Intel",
        "主板型号": "H410M",
        "主板品牌": "Gigabyte",
        "显卡型号": "集成显卡",
        "显卡品牌": "Intel",
    },
    {
        "使用人": "赵六",
        "部门": "财务部",
        "设备名称": "赵六的电脑主机",
        "设备分类": "电脑主机",
        "设备分类编号": "PC",
        "CPU型号": "Intel Core i5-10400",
        "CPU品牌": "Intel",
        "内存型号": "16GB DDR4 2666MHz",
        "内存品牌": "Kingston",
        "硬盘型号": "256GB SSD",
        "硬盘品牌": "Intel",
        "主板型号": "H410M",
        "主板品牌": "Gigabyte",
        "显卡型号": "集成显卡",
        "显卡品牌": "Intel",
    },
    {
        "使用人": "孙七",
        "部门": "市场部",
        "设备名称": "孙七的电脑主机",
        "设备分类": "电脑主机",
        "设备分类编号": "PC",
        "CPU型号": "Intel Core i7-12700",
        "CPU品牌": "Intel",
        "内存型号": "32GB DDR4 3200MHz",
        "内存品牌": "Kingston",
        "硬盘型号": "1TB NVMe SSD",
        "硬盘品牌": "Samsung",
        "主板型号": "B660M",
        "主板品牌": "ASUS",
        "显卡型号": "NVIDIA RTX 3060",
        "显卡品牌": "NVIDIA",
    },
    {
        "使用人": "周八",
        "部门": "市场部",
        "设备名称": "周八的电脑主机",
        "设备分类": "电脑主机",
        "设备分类编号": "PC",
        "CPU型号": "AMD Ryzen 7 5800X",
        "CPU品牌": "AMD",
        "内存型号": "32GB DDR4 3600MHz",
        "内存品牌": "Corsair",
        "硬盘型号": "1TB NVMe SSD",
        "硬盘品牌": "Samsung",
        "主板型号": "B550",
        "主板品牌": "ASUS",
        "显卡型号": "AMD Radeon RX 6700 XT",
        "显卡品牌": "AMD",
    },
]

headers = [
    "使用人", "部门", "设备名称", "设备分类", "设备分类编号",
    "CPU型号", "CPU品牌", "内存型号", "内存品牌", "硬盘型号",
    "硬盘品牌", "主板型号", "主板品牌", "显卡型号", "显卡品牌",
]

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_alignment = Alignment(vertical="center", wrap_text=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "设备导入"

# 表头
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 数据
for row_idx, row_data in enumerate(test_data, 2):
    values = [
        row_data["使用人"], row_data["部门"], row_data["设备名称"],
        row_data["设备分类"], row_data["设备分类编号"],
        row_data["CPU型号"], row_data["CPU品牌"],
        row_data["内存型号"], row_data["内存品牌"],
        row_data["硬盘型号"], row_data["硬盘品牌"],
        row_data["主板型号"], row_data["主板品牌"],
        row_data["显卡型号"], row_data["显卡品牌"],
    ]
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = data_alignment
        cell.border = thin_border

col_widths = [12, 10, 22, 10, 12, 35, 15, 22, 15, 22, 15, 18, 15, 30, 15]
for col, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

ws.row_dimensions[1].height = 25
for row_idx in range(2, len(test_data) + 2):
    ws.row_dimensions[row_idx].height = 22

output_path = "测试导入_6人3部门.xlsx"
wb.save(output_path)
print(f"测试文件已生成: {output_path}")
print(f"共 {len(test_data)} 行数据，3个部门（技术部/财务部/市场部）")
